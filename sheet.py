"""Build the workbook that actually gets emailed to you.

The master schema (email-automation/master_registry.py) is a storage format: every
scraper appends to it, so it is shared and must not churn. This module is the
opposite end — a presentation layer over one run's rows, and the only file that
decides what the sheet you open in Outlook looks like.

Three things it does that the raw master schema doesn't:

1. Source Link, second column. Intent leads come from a public post (a Reddit
   thread, a Freelancer project, an OSM listing). The permalink used to be buried
   mid-sentence in the bullets column, so checking a lead meant hunting for a URL
   inside a paragraph. It gets its own clickable column, immediately right of the
   name.

2. Contact-order sort. Rows are ordered by how easily you can actually reach the
   person: direct email, then Reddit (DM-able), then a phone number, then the
   rest. The top of the sheet is the part worth your time.

3. Widths that fit. Every column is either sized to its real content or wrapped
   at a fixed width with the row grown to match, so nothing is clipped and no
   column border ever needs dragging.
"""

import re

SOURCE_LINK_COLUMN = "Source Link"

# Columns the EMAILED sheet must never carry. They exist in the master schema
# as storage/state (dedup, opt-out flags, drafting inputs) and email-automation
# still needs them while drafting, so they are stripped only by the final
# rewrite that happens right before the sheet is mailed (orchestrator calls
# final_header). The user reads none of them.
DROP_COLUMNS = {
    "Reached", "Do Not Contact", "Channel", "First Reported At",
    "Lead Type", "Has_Website", "Rating", "Reviews",
}


def final_header(header: list) -> list:
    """The stored header minus the columns the emailed sheet must not carry."""
    return [c for c in header if c not in DROP_COLUMNS]

# Ease of contact, best first. The sheet is sorted by these and nothing else
# reorders it, so this tuple IS the priority you read top-down.
TIER_EMAIL = 0     # a direct email address — just write to them
TIER_REDDIT = 1    # a Reddit post — DM the author
TIER_PHONE = 2     # a phone number — call or WhatsApp
TIER_OTHER = 3     # Instagram, a Freelancer bid, or nothing at all

_REDDIT_RE = re.compile(r"reddit\.com", re.I)
_URL_RE = re.compile(r"https?://\S+")
# The Source Link cell is written as =HYPERLINK("url","label") — see _link_cell.
# This reads the url back out, which is what lets the sheet survive the
# write -> draft -> re-read -> rewrite round trip without losing its links.
_HYPERLINK_RE = re.compile(r'^=HYPERLINK\("([^"]+)"', re.I)

# Columns that hold prose. These get a fixed width and wrap; the row is grown to
# fit them. Everything else is auto-sized to its longest actual value, which is
# what stops the sheet arriving with clipped cells.
_WRAP_WIDTHS = {
    "Business Name": 30,
    "Category": 20,
    "Message": 68,
}
_BULLETS_WIDTH = 52          # matched by name: the bullets header is a sentence
_AUTOFIT_MIN = 8
_AUTOFIT_MAX = 44
_AUTOFIT_PAD = 3

_LABEL_MAX = 38              # a Source Link label longer than this is unreadable
_ROW_HEIGHT = 15             # one line of text at the default font
_MAX_ROW_LINES = 12          # cap: a 40-line draft must not own the whole screen


def _is_bullets(col: str) -> bool:
    return col.startswith("Bullet points")


def _wrap_width(col: str):
    """Fixed wrap width for a prose column, or None if it should auto-size."""
    if _is_bullets(col):
        return _BULLETS_WIDTH
    return _WRAP_WIDTHS.get(col)


def _text(row: dict, col: str) -> str:
    """Cell as clean text. openpyxl hands back None for blanks, and str(None) is
    the literal 'None' that was showing up all over the emailed sheet."""
    value = row.get(col)
    if value is None:
        return ""
    return str(value).strip()


def source_link(row: dict) -> str:
    """The public page this lead came from — the one you'd open to check them out
    yourself before writing.

    Best evidence first:
      1. the leeds permalink (the Reddit thread, the Freelancer project, the OSM
         node), which adapter hands over in _source_link;
      2. the same link read back off a sheet we already wrote;
      3. a URL left in the bullets prose, which is where the permalink lived
         before this column existed;
      4. the Instagram profile — scraped businesses have no post to link to, but
         for an Instagram-only business the profile IS the listing you'd check.
    """
    explicit = _text(row, "_source_link") or _text(row, SOURCE_LINK_COLUMN)
    if explicit:
        formula = _HYPERLINK_RE.match(explicit)
        if formula:
            return formula.group(1)
        if explicit.lower().startswith("http"):
            return explicit
        # Anything else in this column is a bare label from a hand-edited sheet:
        # not a link, so keep looking rather than emitting a broken one.
    for col in row:
        if _is_bullets(col):
            match = _URL_RE.search(_text(row, col))
            if match:
                return match.group(0).rstrip(").,")
    handle = _text(row, "Instagram").lstrip("@")
    if handle:
        if handle.startswith("http"):
            return handle
        return f"https://www.instagram.com/{handle}/"
    return ""


def link_label(url: str) -> str:
    """What the Source Link cell shows. The cell is a real hyperlink, so this is
    only what you READ — one click still opens the URL itself.

    A Reddit permalink runs to ~116 characters, and no column width makes that
    scannable; pasting it in raw is what forces the reader to drag borders. So
    show the part that tells you where you're going ("reddit.com/r/smallbusiness")
    and let the click carry the rest.
    """
    if not url:
        return ""
    bare = re.sub(r"^https?://(www\.)?", "", url).rstrip("/")
    host = bare.split("/", 1)[0].lower()
    parts = bare.split("/")

    if "reddit.com" in host and len(parts) >= 3 and parts[1] == "r":
        label = f"reddit.com/r/{parts[2]}"           # the subreddit is the signal
    elif "freelancer" in host and len(parts) >= 3:
        # .../projects/<category>/<project-slug>. The category repeats across
        # every job ("web-development" twice over tells you nothing), so name the
        # project when it's there and keep the category only as a fallback.
        slug = parts[3] if len(parts) >= 4 else parts[2]
        label = f"freelancer.com/{re.sub(r'-\d+$', '', slug)}"
    elif "openstreetmap.org" in host:
        label = bare                                 # already short: .../node/123
    elif "instagram.com" in host and len(parts) >= 2:
        label = f"instagram.com/{parts[1]}"
    else:
        label = bare
    if len(label) > _LABEL_MAX:
        label = label[:_LABEL_MAX - 1] + "…"
    return label


def _link_cell(cell, url: str) -> None:
    """Show the readable label, open the real URL, and keep the URL in the cell
    VALUE rather than only in openpyxl's link metadata.

    That last part is load-bearing. The orchestrator writes this sheet, hands it
    to email-automation to draft into, then reads it back and rewrites it —
    and a read-back drops everything except values. A =HYPERLINK() formula is
    the one form that carries the url, the label and the click through that
    round trip intact.
    """
    from openpyxl.styles import Alignment

    label = link_label(url)
    if '"' in url:                     # a quote would end the formula string
        cell.value = url
    else:
        cell.value = f'=HYPERLINK("{url}","{label}")'
    cell.style = "Hyperlink"
    # A named style carries its own alignment, so it silently undoes the one set
    # when the row was written — which left every link floating at the bottom of
    # a tall row, nowhere near the business it belongs to. Re-apply it last.
    cell.alignment = Alignment(vertical="top")


def contact_tier(row: dict) -> int:
    """How easily this lead can be reached. Lower sorts first."""
    if _text(row, "Email Address"):
        return TIER_EMAIL
    if _REDDIT_RE.search(source_link(row)):
        return TIER_REDDIT
    if _text(row, "Phone Number"):
        return TIER_PHONE
    return TIER_OTHER


def sort_rows(rows: list[dict]) -> list[dict]:
    """Contact tier first, then business name so each tier is stable and
    scannable rather than shuffled by scrape order."""
    return sorted(rows, key=lambda r: (contact_tier(r),
                                       _text(r, "Business Name").lower()))


def build_header(header: list) -> list:
    """The stored header with Source Link forced into second position."""
    out = [c for c in header if c != SOURCE_LINK_COLUMN]
    at = 1 if out and out[0] == "Business Name" else 0
    out.insert(at, SOURCE_LINK_COLUMN)
    return out


def _row_height(values: list, header: list) -> float:
    """Tall enough for the tallest wrapped cell in the row, capped. Excel does
    not auto-fit rows that openpyxl wrote, so leaving the height unset renders a
    3-line bullet list as one clipped line."""
    lines = 1
    for col, value in zip(header, values):
        width = _wrap_width(col)
        if not width or not value:
            continue
        used = sum(max(1, -(-len(para) // width))      # ceil: wrapped line count
                   for para in str(value).split("\n"))
        lines = max(lines, used)
    return min(lines, _MAX_ROW_LINES) * _ROW_HEIGHT


def write(path: str, header: list, rows: list) -> None:
    """Write the emailed workbook: Source Link second, contact-order sort, and
    formatting that is readable the moment it opens."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header = build_header(header)
    rows = sort_rows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    ws.append(header)
    head_fill = PatternFill("solid", fgColor="1F2430")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center", horizontal="left",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 30

    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    # Longest value seen per column, for the auto-sized ones. The header itself
    # is the floor: a column must at least show its own name.
    longest = {col: len(col) for col in header}

    for row in rows:
        link = source_link(row)
        values = [link_label(link) if col == SOURCE_LINK_COLUMN else _text(row, col)
                  for col in header]
        ws.append(values)
        excel_row = ws.max_row

        for i, (col, value) in enumerate(zip(header, values), 1):
            cell = ws.cell(row=excel_row, column=i)
            cell.alignment = wrap if _wrap_width(col) else top
            if not _wrap_width(col) and value:
                longest[col] = max(longest[col],
                                   max(len(p) for p in value.split("\n")))
        if link:
            _link_cell(ws.cell(row=excel_row,
                               column=header.index(SOURCE_LINK_COLUMN) + 1), link)
        ws.row_dimensions[excel_row].height = _row_height(values, header)

    for i, col in enumerate(header, 1):
        fixed = _wrap_width(col)
        width = fixed if fixed else min(max(longest[col] + _AUTOFIT_PAD,
                                            _AUTOFIT_MIN), _AUTOFIT_MAX)
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze the header and the business name, so scrolling right through the
    # long prose columns keeps showing you which lead you are looking at.
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    wb.save(path)
