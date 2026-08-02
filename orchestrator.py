"""Per-button pipelines: run the tool, isolate this run's new leads, draft them
with email-automation, and email the sheet to you.

Nothing here contacts a lead. Scrapers/leeds run with --no-email (or never email),
email-automation only ever drafts, and the single outbound email goes to you.
"""

import os
import subprocess
import sys

import config
import adapter
import mailer
import ratelimit
import sheet

sys.path.insert(0, config.EMAIL_AUTOMATION_DIR)
import master_registry as mr   # email-automation's schema/identity helpers


# --- xlsx helpers (header-name based, position independent) ------------------

def _read_rows(path: str):
    """(header, [row_dict, ...]). ([], []) if the file is absent."""
    if not os.path.exists(path):
        return [], []
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(h) for h in (next(it, None) or []) if h is not None]
    rows = []
    for values in it:
        if not any(v is not None for v in values):
            continue
        rows.append({col: (values[i] if i < len(values) and values[i] is not None else "")
                     for i, col in enumerate(header)})
    wb.close()
    return header, rows


def _write_rows(path: str, header: list, rows: list):
    """The emailed sheet. sheet.py owns every layout decision — Source Link in
    column B, contact-order sort, widths — so this stays a one-liner."""
    sheet.write(path, header, rows)


# --- subprocess runner -------------------------------------------------------

def _run(python, args, cwd, env_extra, repo_hint, log, hits):
    """Stream a subprocess, mirroring every line to the live log and scanning it
    for rate-limit signals. Returns the exit code."""
    env = os.environ.copy()
    env["PYTHONUNBUFFERED"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    env.update(env_extra or {})

    log(f"$ {os.path.basename(python)} {' '.join(args)}   (cwd={os.path.basename(cwd)})")
    proc = subprocess.Popen(
        [python] + args, cwd=cwd, env=env,
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        stdin=subprocess.DEVNULL,
        text=True, encoding="utf-8", errors="replace", bufsize=1,
    )
    for line in iter(proc.stdout.readline, ""):
        line = line.rstrip("\n")
        if line:
            log(line)
            for hit in ratelimit.detect(line, repo_hint):
                sig = (hit["provider"], hit["env_path"], hit["env_key"])
                if sig not in {(h["provider"], h["env_path"], h["env_key"]) for h in hits}:
                    hits.append(hit)
    proc.stdout.close()
    return proc.wait()


def _draft(temp_file, log, hits):
    """Run email-automation --prep against temp_file (fills Message/Channel)."""
    ea = config.EMAIL_AUTOMATION
    code = _run(ea["python"], ea["args"], ea["cwd"],
                {"MASTER_FILE": temp_file}, "email-automation", log, hits)
    if code != 0:
        log(f"[launcher] email-automation exited {code} (drafts may be partial)")


def _mail(temp_file, label, count, drafted, log, hits, body=None):
    subject = f"Chillispark leads, {label}: {count} new"
    body = body or (
        f"<p><b>{count}</b> new lead(s) from <b>{label}</b>, "
        f"<b>{drafted}</b> with a ready cold-outreach draft in the Message column.</p>"
        f"<p>Sorted by how easily you can reach them: direct email first, then "
        f"Reddit, then phone numbers, then the rest. Column B links straight to "
        f"the original post so you can check any lead yourself.</p>"
        f"<p>Open the attached sheet, review each draft, then send by hand.</p>"
    )
    ok, status, detail = mailer.send(temp_file, subject, body)
    log(f"[launcher] mail: {detail}")
    if not ok:
        for hit in ratelimit.detect(detail, "launcher"):
            hits.append(hit)
    return ok, detail


# --- pipelines ---------------------------------------------------------------

def _pipeline_scraper(button_id, btn, jobid, log, hits):
    os.makedirs(config.OUT_DIR, exist_ok=True)
    master_run = config.MASTER_FILE

    _, before = _read_rows(master_run)
    keys_before = set()
    for r in before:
        keys_before |= mr.identity_keys(r)
    log(f"[launcher] master had {len(before)} leads before this run")

    code = _run(btn["python"], btn["args"], btn["cwd"],
                {"MASTER_FILE": master_run}, btn["env_repo"], log, hits)
    log(f"[launcher] {btn['label']} exited {code}")

    header, after = _read_rows(master_run)
    new_rows = [r for r in after if not (mr.identity_keys(r) & keys_before)]
    log(f"[launcher] {len(new_rows)} new lead(s) this run")
    if not new_rows:
        return {"count": 0, "drafted": 0, "mailed": False, "detail": "no new leads"}

    temp = os.path.join(config.OUT_DIR, f"run_{button_id}_{jobid}.xlsx")
    _write_rows(temp, header, new_rows)
    _draft(temp, log, hits)

    # Final rewrite: drop the state/noise columns the reader never uses. Must
    # happen AFTER drafting — email-automation still reads some of them.
    header, drafted_rows = _read_rows(temp)
    _write_rows(temp, sheet.final_header(header), drafted_rows)
    drafted = sum(1 for r in drafted_rows if str(r.get("Message") or "").strip())
    ok, detail = _mail(temp, btn["label"], len(new_rows), drafted, log, hits)
    return {"count": len(new_rows), "drafted": drafted, "mailed": ok,
            "detail": detail, "file": temp}


def _pipeline_leeds(button_id, btn, jobid, log, hits):
    os.makedirs(config.OUT_DIR, exist_ok=True)
    source = btn["source_file"]

    _, before = _read_rows(source)
    before_links = {str(r.get("Permalink") or "").strip()
                    for r in before if str(r.get("Permalink") or "").strip()}
    log(f"[launcher] intent sheet had {len(before)} rows before this run")

    code = _run(btn["python"], btn["args"], btn["cwd"], {}, btn["env_repo"], log, hits)
    log(f"[launcher] {btn['label']} exited {code}")

    _, after = _read_rows(source)
    if before_links:
        new_intent = [r for r in after
                      if str(r.get("Permalink") or "").strip() not in before_links]
    else:
        new_intent = after   # first ever run — everything is new
    log(f"[launcher] {len(new_intent)} new intent lead(s) this run")

    new_intent, foreign = adapter.contactable(new_intent)
    if foreign:
        log(f"[launcher] dropped {foreign} OSM business(es) outside India — "
            f"no way to contact them")
    if not new_intent:
        return {"count": 0, "drafted": 0, "mailed": False, "detail": "no new leads"}

    master_rows = adapter.convert(new_intent)
    opener_map = {}
    for m in master_rows:
        for k in m.pop("_keys", set()):
            opener_map[k] = m.get("_opener", "")
        m.pop("_opener", None)

    temp = os.path.join(config.OUT_DIR, f"run_{button_id}_{jobid}.xlsx")
    _write_rows(temp, mr.COLUMNS, master_rows)
    _draft(temp, log, hits)

    # Fallback: where email-automation had no channel to draft, keep leeds' opener.
    header, drafted_rows = _read_rows(temp)
    filled = 0
    for r in drafted_rows:
        if not str(r.get("Message") or "").strip():
            for k in mr.identity_keys(r):
                if opener_map.get(k):
                    r["Message"] = opener_map[k]
                    filled += 1
                    break
    # This rewrite is also the final one, so the state/noise columns go here.
    _write_rows(temp, sheet.final_header(header), drafted_rows)
    drafted = sum(1 for r in drafted_rows if str(r.get("Message") or "").strip())
    log(f"[launcher] drafts: {drafted} total ({filled} filled from leeds' opener)")

    ok, detail = _mail(temp, btn["label"], len(new_intent), drafted, log, hits)
    return {"count": len(new_intent), "drafted": drafted, "mailed": ok,
            "detail": detail, "file": temp}


def _pipeline_dm(button_id, btn, jobid, log, hits, category):
    """Run scraper3's DM mode for one chosen category. scraper3 writes the
    10-row sheet of clickable Instagram profile + ig.me DM links straight to
    temp (and records the picks in the master so the next run is fresh); the
    launcher only mails it. No drafting step — the sheet IS the work list."""
    os.makedirs(config.OUT_DIR, exist_ok=True)
    temp = os.path.join(config.OUT_DIR, f"run_{button_id}_{jobid}.xlsx")

    args = ["main.py", "--category-name", category, "--dm-out", temp, "--no-email"]
    code = _run(btn["python"], args, btn["cwd"], {}, btn["env_repo"], log, hits)
    log(f"[launcher] {btn['label']} exited {code}")

    _, rows = _read_rows(temp)
    if not rows:
        return {"count": 0, "drafted": 0, "mailed": False, "detail": "no new leads"}

    label = f"{btn['label']} - {category}"
    body = (
        f"<p><b>{len(rows)}</b> Instagram-only <b>{category}</b> business(es), "
        f"fresh this run.</p>"
        f"<p>The DM column opens their Instagram chat directly (ig.me); the "
        f"Profile column opens the account so you can vet it in ten seconds "
        f"first. Send by hand, 10-20 a day, no links in the first message.</p>"
    )
    ok, detail = _mail(temp, label, len(rows), 0, log, hits, body=body)
    return {"count": len(rows), "drafted": 0, "mailed": ok,
            "detail": detail, "file": temp}


def _pipeline_niche(button_id, btn, jobid, log, hits, niche):
    """Run the Niche tool for one chosen niche. Niche writes the finished sheet
    (one-liner Message included, its own seen.json dedup) straight to temp; the
    launcher only mails it. No email-automation drafting step — the one-liner
    is the message."""
    os.makedirs(config.OUT_DIR, exist_ok=True)
    temp = os.path.join(config.OUT_DIR, f"run_{button_id}_{jobid}.xlsx")

    args = ["main.py", "--niche", niche, "--no-email", "--out", temp]
    code = _run(btn["python"], args, btn["cwd"], {}, btn["env_repo"], log, hits)
    log(f"[launcher] {btn['label']} exited {code}")

    _, rows = _read_rows(temp)
    if not rows:
        return {"count": 0, "drafted": 0, "mailed": False, "detail": "no new leads"}

    drafted = sum(1 for r in rows if str(r.get("Message") or "").strip())
    label = f"{btn['label']} - {niche}"
    body = (
        f"<p><b>{len(rows)}</b> new <b>{niche}</b> business(es) with no website, "
        f"each with a ready one-line pitch in the Message column.</p>"
        f"<p>Sorted easiest-to-reach first; the WhatsApp column links straight "
        f"to a chat. Copy the Message, send by hand.</p>"
    )
    ok, detail = _mail(temp, label, len(rows), drafted, log, hits, body=body)
    return {"count": len(rows), "drafted": drafted, "mailed": ok,
            "detail": detail, "file": temp}


def run(button_id: str, jobid: str, log, params=None):
    """Entry point used by the server's background thread. Returns a result dict
    including any rate-limit hits gathered along the way. `params` carries
    per-run UI choices (today: the Niche button's selected niche)."""
    btn = config.BUTTONS[button_id]
    hits: list[dict] = []
    log(f"[launcher] === {btn['label']} ({btn['subtitle']}) ===")
    log(f"[launcher] recipient={config.RECIPIENT}")
    try:
        choice = str((params or {}).get("choice") or "").strip()
        if btn["kind"] == "scraper":
            result = _pipeline_scraper(button_id, btn, jobid, log, hits)
        elif btn["kind"] == "dm":
            if not choice:
                raise ValueError("no category selected")
            result = _pipeline_dm(button_id, btn, jobid, log, hits, choice)
        elif btn["kind"] == "niche":
            if not choice:
                raise ValueError("no niche selected")
            result = _pipeline_niche(button_id, btn, jobid, log, hits, choice)
        else:
            result = _pipeline_leeds(button_id, btn, jobid, log, hits)
        result["ok"] = True
    except Exception as exc:
        import traceback
        log("[launcher] ERROR: " + "".join(traceback.format_exc()))
        result = {"ok": False, "count": 0, "drafted": 0, "mailed": False,
                  "detail": f"{type(exc).__name__}: {exc}"}
    result["rate_limits"] = hits
    result["label"] = btn["label"]
    result["recipient"] = config.RECIPIENT
    return result
